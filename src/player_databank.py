"""Player Databank: game log fetching, loading, and stat view configuration.

Provides per-game stat storage, retrieval, and the STAT_VIEW_OPTIONS /
STAT_VIEW_PARAMS dictionaries that drive the Player Databank UI dropdowns.
"""

import logging
import math
from datetime import UTC, datetime, timedelta
from io import BytesIO

import pandas as pd

from src.database import get_connection, load_player_pool
from src.ui_shared import T

logger = logging.getLogger(__name__)

try:
    import statsapi
except ImportError:
    statsapi = None  # type: ignore[assignment]

# ── Fantasy category column groups ───────────────────────────────────────────

HITTING_CATS = ["R", "HR", "RBI", "SB", "AVG", "OBP"]
PITCHING_CATS = ["W", "L", "SV", "K", "ERA", "WHIP"]

HITTING_COLS_TOTAL = ["h", "ab", "r", "hr", "rbi", "sb", "pa", "bb", "hbp", "sf"]
PITCHING_COLS_TOTAL = ["ip", "w", "l", "sv", "k", "er", "bb_allowed", "h_allowed"]

# ── Stat view options (28 keys) ───────────────────────────────────────────────

STAT_VIEW_OPTIONS: dict[str, str] = {
    "S_PS7": "Next 7 Days (proj)",
    "S_PS14": "Next 14 Days (proj)",
    "S_PSR": "Remaining Games (proj)",
    "S_L": "Today (live)",
    "S_L7": "Last 7 Days (total)",
    "S_L14": "Last 14 Days (total)",
    "S_L30": "Last 30 Days (total)",
    "S_S_2026": "Season (total)",
    "S_S_2025": "2025 Season (total)",
    "S_S_2024": "2024 Season (total)",
    "ADVST_ADVST_2026": "2026 Advanced",
    "ADVST_ADVST_2025": "2025 Advanced",
    "S_AL7": "Last 7 Days (avg)",
    "S_AL14": "Last 14 Days (avg)",
    "S_AL30": "Last 30 Days (avg)",
    "S_AS_2026": "Season (avg)",
    "S_AS_2025": "2025 Season (avg)",
    "S_AS_2024": "2024 Season (avg)",
    "S_SDL7": "Last 7 Days (std dev)",
    "S_SDL14": "Last 14 Days (std dev)",
    "S_SDL30": "Last 30 Days (std dev)",
    "S_SD_2026": "Season (std dev)",
    "S_SD_2025": "2025 Season (std dev)",
    "S_SD_2024": "2024 Season (std dev)",
    "K_K": "Ranks",
    "R_O": "Research",
    "M_W": "Fantasy Matchups",
    "O_O": "Opponents",
}

# ── Stat view computation parameters ─────────────────────────────────────────
# Each entry maps a STAT_VIEW_OPTIONS key to its computation parameters:
#   type  : "proj" | "live" | "total" | "avg" | "stddev" | "special"
#   days  : number of days back (None = entire season or special handling)
#   season: integer season year (None = current / multi-season)

STAT_VIEW_PARAMS: dict[str, dict] = {
    "S_PS7": {"type": "proj", "days": 7, "season": None},
    "S_PS14": {"type": "proj", "days": 14, "season": None},
    "S_PSR": {"type": "proj", "days": None, "season": None},
    "S_L": {"type": "live", "days": 0, "season": None},
    "S_L7": {"type": "total", "days": 7, "season": None},
    "S_L14": {"type": "total", "days": 14, "season": None},
    "S_L30": {"type": "total", "days": 30, "season": None},
    "S_S_2026": {"type": "total", "days": None, "season": 2026},
    "S_S_2025": {"type": "total", "days": None, "season": 2025},
    "S_S_2024": {"type": "total", "days": None, "season": 2024},
    "ADVST_ADVST_2026": {"type": "advanced", "days": None, "season": 2026},
    "ADVST_ADVST_2025": {"type": "advanced", "days": None, "season": 2025},
    "S_AL7": {"type": "avg", "days": 7, "season": None},
    "S_AL14": {"type": "avg", "days": 14, "season": None},
    "S_AL30": {"type": "avg", "days": 30, "season": None},
    "S_AS_2026": {"type": "avg", "days": None, "season": 2026},
    "S_AS_2025": {"type": "avg", "days": None, "season": 2025},
    "S_AS_2024": {"type": "avg", "days": None, "season": 2024},
    "S_SDL7": {"type": "stddev", "days": 7, "season": None},
    "S_SDL14": {"type": "stddev", "days": 14, "season": None},
    "S_SDL30": {"type": "stddev", "days": 30, "season": None},
    "S_SD_2026": {"type": "stddev", "days": None, "season": 2026},
    "S_SD_2025": {"type": "stddev", "days": None, "season": 2025},
    "S_SD_2024": {"type": "stddev", "days": None, "season": 2024},
    "K_K": {"type": "special", "days": None, "season": None},
    "R_O": {"type": "special", "days": None, "season": None},
    "M_W": {"type": "special", "days": None, "season": None},
    "O_O": {"type": "special", "days": None, "season": None},
}


# ── Game log loading ──────────────────────────────────────────────────────────


def load_game_logs(
    player_ids: list[int],
    days: int | None = None,
    season: int | None = None,
) -> pd.DataFrame:
    """Load game logs from the local SQLite database.

    Args:
        player_ids: List of player IDs to query.
        days: If provided, restrict to games within the last N days from today.
              If None, no date-range filter is applied (use ``season`` to
              restrict by year instead).
        season: If provided, restrict to games for this season year.
                If None, no season filter is applied.

    Returns:
        DataFrame of matching game log rows. Empty DataFrame if no matches or
        if ``player_ids`` is empty.
    """
    if not player_ids:
        return pd.DataFrame()

    conn = get_connection()
    try:
        placeholders = ",".join("?" * len(player_ids))
        params: list = list(player_ids)
        where_clauses = [f"player_id IN ({placeholders})"]

        if season is not None:
            where_clauses.append("season = ?")
            params.append(season)

        if days is not None:
            cutoff = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%d")
            where_clauses.append("game_date >= ?")
            params.append(cutoff)

        where_sql = " AND ".join(where_clauses)
        sql = f"SELECT * FROM game_logs WHERE {where_sql} ORDER BY game_date DESC"

        df = pd.read_sql_query(sql, conn, params=params)
        return df
    except Exception:
        logger.exception("Failed to load game logs for player_ids=%s", player_ids)
        return pd.DataFrame()
    finally:
        conn.close()


# ── Rolling stat computation ──────────────────────────────────────────────────


def compute_rolling_stats(
    player_ids: list[int],
    days: int | None = None,
    stat_type: str = "total",
    season: int = 2026,
) -> pd.DataFrame:
    """Compute rolling window stats from game logs.

    Args:
        player_ids: Player IDs to compute for.
        days: Number of days in the rolling window. None = entire season.
        stat_type: "total", "avg", or "stddev".
        season: Season year.

    Returns:
        DataFrame with one row per player, stat columns computed per stat_type.
        For "total" and "avg", computed rate stat columns are appended:
        ``avg_calc``, ``obp_calc``, ``era_calc``, ``whip_calc``.
    """
    logs = load_game_logs(player_ids, days=days, season=season)
    if logs.empty:
        return pd.DataFrame()

    # Determine numeric columns present in the result
    all_count_cols = HITTING_COLS_TOTAL + PITCHING_COLS_TOTAL
    num_cols = [c for c in all_count_cols if c in logs.columns]

    if stat_type == "stddev":
        result = logs.groupby("player_id")[num_cols].std(ddof=1).fillna(0).reset_index()
        return result

    # "total" and "avg" both start from sums
    sums = logs.groupby("player_id")[num_cols].sum()
    game_counts = logs.groupby("player_id").size().rename("_game_count")
    result = sums.join(game_counts).reset_index()

    if stat_type == "avg":
        for col in num_cols:
            result[col] = result[col] / result["_game_count"]

    # Rename to games_played so the databank table can show GP
    result = result.rename(columns={"_game_count": "games_played"})

    # Append computed rate stats (using weighted formulas per CLAUDE.md)
    def _safe_div(num: pd.Series, den: pd.Series) -> pd.Series:
        """Divide num by den, returning NaN where denominator is zero."""
        return num.where(den != 0).div(den.where(den != 0))

    # AVG = sum(h) / sum(ab)
    if "h" in result.columns and "ab" in result.columns:
        result["avg_calc"] = _safe_div(result["h"], result["ab"])

    # OBP = sum(h + bb + hbp) / sum(ab + bb + hbp + sf)
    if all(c in result.columns for c in ["h", "bb", "hbp", "ab", "sf"]):
        obp_num = result["h"] + result["bb"] + result["hbp"]
        obp_den = result["ab"] + result["bb"] + result["hbp"] + result["sf"]
        result["obp_calc"] = _safe_div(obp_num, obp_den)

    # ERA = sum(er) * 9 / sum(ip)
    if "er" in result.columns and "ip" in result.columns:
        result["era_calc"] = _safe_div(result["er"] * 9, result["ip"])

    # WHIP = sum(bb_allowed + h_allowed) / sum(ip)
    if all(c in result.columns for c in ["bb_allowed", "h_allowed", "ip"]):
        whip_num = result["bb_allowed"] + result["h_allowed"]
        result["whip_calc"] = _safe_div(whip_num, result["ip"])

    return result


# ── Game log fetching from MLB Stats API ──────────────────────────────────────


def fetch_game_logs_from_api(
    season: int = 2026,
    force: bool = False,
) -> int:
    """Fetch per-game stats from the MLB Stats API and persist to SQLite.

    Iterates through all players in the ``players`` table that have a non-null
    ``mlb_id``, calls ``statsapi.player_stat_data()`` for each, and stores the
    results into the ``game_logs`` table via INSERT OR REPLACE.

    Args:
        season: The MLB season year to fetch game logs for.
        force: If True, re-fetch even if records already exist. (Currently
               INSERT OR REPLACE always overwrites; this flag is reserved for
               future staleness checks.)

    Returns:
        Total number of game-log rows inserted / replaced across all players.
    """
    if statsapi is None:
        logger.warning("statsapi not installed — skipping fetch_game_logs_from_api")
        return 0

    # Load all players with a known MLB ID from the DB
    conn = get_connection()
    try:
        players_df = pd.read_sql_query(
            "SELECT player_id, mlb_id, is_hitter FROM players WHERE mlb_id IS NOT NULL",
            conn,
        )
    except Exception:
        logger.exception("Failed to read players table for game log fetch")
        return 0
    finally:
        conn.close()

    if players_df.empty:
        logger.info("No players with mlb_id found; skipping game log fetch")
        return 0

    total_inserted = 0

    for _, row in players_df.iterrows():
        player_id = int(row["player_id"])
        mlb_id = int(row["mlb_id"])
        is_hitter = int(row.get("is_hitter", 1))

        try:
            inserted = _fetch_and_store_player_logs(
                player_id=player_id,
                mlb_id=mlb_id,
                is_hitter=is_hitter,
                season=season,
            )
            total_inserted += inserted
        except Exception:
            logger.warning(
                "Error fetching game logs for player_id=%d mlb_id=%d",
                player_id,
                mlb_id,
                exc_info=True,
            )

    logger.info(
        "fetch_game_logs_from_api: inserted/replaced %d rows for season %d",
        total_inserted,
        season,
    )
    return total_inserted


def _fetch_and_store_player_logs(
    player_id: int,
    mlb_id: int,
    is_hitter: int,
    season: int,
) -> int:
    """Fetch game logs for a single player and write to game_logs table.

    Tries the appropriate stat group first (hitting for hitters, pitching for
    pitchers). Falls back to the other group if the primary returns no data.

    Args:
        player_id: Internal DB player ID (FK into ``players`` table).
        mlb_id: MLB Stats API player ID.
        is_hitter: 1 if the player is primarily a hitter, 0 for pitcher.
        season: Season year.

    Returns:
        Number of rows inserted.
    """
    groups = ["hitting", "pitching"] if is_hitter else ["pitching", "hitting"]
    rows: list[dict] = []

    for group in groups:
        try:
            result = statsapi.player_stat_data(
                mlb_id,
                group=group,
                type="gameLog",
                sportId=1,
            )
            entries = result.get("stats", [])
            if not entries:
                continue

            for entry in entries:
                game_date = entry.get("date", "")
                if not game_date:
                    continue

                raw = entry.get("stats", entry)
                row = _parse_game_log_row(player_id, game_date, season, group, raw)
                rows.append(row)

            # If we found data in this group, don't try the other
            break
        except Exception:
            logger.debug("No %s gameLog for mlb_id=%d", group, mlb_id, exc_info=False)

    if not rows:
        return 0

    conn = get_connection()
    try:
        conn.executemany(
            """
            INSERT OR REPLACE INTO game_logs
                (player_id, game_date, season,
                 pa, ab, h, r, hr, rbi, sb, bb, hbp, sf,
                 ip, w, l, sv, k, er, bb_allowed, h_allowed)
            VALUES
                (:player_id, :game_date, :season,
                 :pa, :ab, :h, :r, :hr, :rbi, :sb, :bb, :hbp, :sf,
                 :ip, :w, :l, :sv, :k, :er, :bb_allowed, :h_allowed)
            """,
            rows,
        )
        conn.commit()
        return len(rows)
    except Exception:
        logger.exception("Failed to store game logs for player_id=%d", player_id)
        return 0
    finally:
        conn.close()


def _parse_game_log_row(
    player_id: int,
    game_date: str,
    season: int,
    group: str,
    raw: dict,
) -> dict:
    """Convert a raw statsapi game-log entry into a game_logs table row dict.

    Args:
        player_id: Internal DB player ID.
        game_date: ISO date string (YYYY-MM-DD).
        season: Season year integer.
        group: "hitting" or "pitching".
        raw: The ``stats`` sub-dict from the statsapi response entry.

    Returns:
        Dict with all game_logs column keys populated (zero-defaults for
        columns from the other stat group).
    """

    def _int(key: str) -> int:
        try:
            return int(raw.get(key, 0) or 0)
        except (ValueError, TypeError):
            return 0

    def _float(key: str) -> float:
        try:
            val = raw.get(key, 0) or 0
            # IP is sometimes stored as "2.1" (outs notation) but statsapi
            # returns decimal innings directly (e.g. 6.0 for 6 full innings).
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    base: dict = {
        "player_id": player_id,
        "game_date": game_date,
        "season": season,
        # Hitting
        "pa": 0,
        "ab": 0,
        "h": 0,
        "r": 0,
        "hr": 0,
        "rbi": 0,
        "sb": 0,
        "bb": 0,
        "hbp": 0,
        "sf": 0,
        # Pitching
        "ip": 0.0,
        "w": 0,
        "l": 0,
        "sv": 0,
        "k": 0,
        "er": 0,
        "bb_allowed": 0,
        "h_allowed": 0,
    }

    if group == "hitting":
        base.update(
            {
                "pa": _int("plateAppearances"),
                "ab": _int("atBats"),
                "h": _int("hits"),
                "r": _int("runs"),
                "hr": _int("homeRuns"),
                "rbi": _int("rbi"),
                "sb": _int("stolenBases"),
                "bb": _int("baseOnBalls"),
                "hbp": _int("hitByPitch"),
                "sf": _int("sacFlies"),
            }
        )
    else:  # pitching
        base.update(
            {
                "ip": _float("inningsPitched"),
                "w": _int("wins"),
                "l": _int("losses"),
                "sv": _int("saves"),
                "k": _int("strikeOuts"),
                "er": _int("earnedRuns"),
                "bb_allowed": _int("baseOnBalls"),
                "h_allowed": _int("hits"),
            }
        )

    return base


# ── Databank enrichment helpers ──────────────────────────────────────────────

# Timeout for statsapi calls (seconds)
_API_TIMEOUT = 30

# MLB team full name → abbreviation map (for schedule parsing)
_TEAM_ABBR: dict[str, str] = {
    "Arizona Diamondbacks": "ARI",
    "Atlanta Braves": "ATL",
    "Baltimore Orioles": "BAL",
    "Boston Red Sox": "BOS",
    "Chicago Cubs": "CHC",
    "Chicago White Sox": "CHW",
    "Cincinnati Reds": "CIN",
    "Cleveland Guardians": "CLE",
    "Colorado Rockies": "COL",
    "Detroit Tigers": "DET",
    "Houston Astros": "HOU",
    "Kansas City Royals": "KC",
    "Los Angeles Angels": "LAA",
    "Los Angeles Dodgers": "LAD",
    "Miami Marlins": "MIA",
    "Milwaukee Brewers": "MIL",
    "Minnesota Twins": "MIN",
    "New York Mets": "NYM",
    "New York Yankees": "NYY",
    "Athletics": "OAK",
    "Oakland Athletics": "OAK",
    "Philadelphia Phillies": "PHI",
    "Pittsburgh Pirates": "PIT",
    "San Diego Padres": "SD",
    "San Francisco Giants": "SF",
    "Seattle Mariners": "SEA",
    "St. Louis Cardinals": "STL",
    "Tampa Bay Rays": "TB",
    "Texas Rangers": "TEX",
    "Toronto Blue Jays": "TOR",
    "Washington Nationals": "WSH",
}


def get_todays_opponent_map() -> dict[str, str]:
    """Build MLB team abbreviation → today's opponent mapping.

    Returns dict like ``{"NYY": "@ BOS", "BOS": "vs NYY"}``.
    Uses ``get_target_game_date()`` which auto-advances to tomorrow when
    all of today's games are final.  Returns empty dict on error.
    """
    if statsapi is None:
        return {}
    try:
        from src.game_day import get_target_game_date

        target = get_target_game_date()
        date_str = target.strftime("%Y-%m-%d") if hasattr(target, "strftime") else str(target)
        schedule = statsapi.schedule(date=date_str)
    except Exception:
        logger.debug("get_todays_opponent_map: schedule fetch failed", exc_info=True)
        return {}

    opp_map: dict[str, str] = {}
    for game in schedule:
        home = _TEAM_ABBR.get(game.get("home_name", ""), "")
        away = _TEAM_ABBR.get(game.get("away_name", ""), "")
        if home and away:
            opp_map[home] = f"vs {away}"
            opp_map[away] = f"@ {home}"
    return opp_map


def get_adds_drops_map() -> tuple[dict[str, int], dict[str, int]]:
    """Aggregate add/drop counts per player from Yahoo transactions.

    Returns ``(adds_by_name, drops_by_name)`` dicts.
    Returns empty dicts if Yahoo is unavailable.
    """
    try:
        from src.yahoo_data_service import get_yahoo_data_service

        yds = get_yahoo_data_service()
        if not yds.is_connected():
            return {}, {}
        txns = yds.get_transactions()
        if txns is None or txns.empty:
            return {}, {}
    except Exception:
        logger.debug("get_adds_drops_map: Yahoo unavailable", exc_info=True)
        return {}, {}

    adds: dict[str, int] = {}
    drops: dict[str, int] = {}
    name_col = "player_name" if "player_name" in txns.columns else "name"
    type_col = "type" if "type" in txns.columns else "transaction_type"
    if name_col not in txns.columns or type_col not in txns.columns:
        return {}, {}

    for _, row in txns.iterrows():
        pname = str(row.get(name_col, ""))
        ttype = str(row.get(type_col, "")).lower()
        if not pname:
            continue
        if "add" in ttype:
            adds[pname] = adds.get(pname, 0) + 1
        if "drop" in ttype:
            drops[pname] = drops.get(pname, 0) + 1
    return adds, drops


def _enrich_databank_columns(df: pd.DataFrame) -> None:
    """Enrich databank DataFrame in-place with opponent, adds, drops columns."""
    # Opponent
    try:
        opp_map = get_todays_opponent_map()
        if opp_map and "team" in df.columns:
            df["opponent"] = df["team"].map(opp_map).fillna("-")
        else:
            df["opponent"] = "-"
    except Exception:
        df["opponent"] = "-"

    # Adds / Drops
    try:
        adds_map, drops_map = get_adds_drops_map()
        name_col = "player_name" if "player_name" in df.columns else "name"
        if name_col in df.columns:
            df["adds"] = df[name_col].map(adds_map).fillna(0).astype(int)
            df["drops"] = df[name_col].map(drops_map).fillna(0).astype(int)
        else:
            df["adds"] = 0
            df["drops"] = 0
    except Exception:
        df["adds"] = 0
        df["drops"] = 0


# ── Databank assembly ─────────────────────────────────────────────────────────


def load_databank(
    stat_view: str,
    season: int = 2026,
) -> pd.DataFrame:
    """Load a player pool DataFrame for the given stat view key.

    Uses ``STAT_VIEW_PARAMS`` to determine the computation type and dispatches
    to the appropriate data source.  The returned DataFrame always has a
    ``player_name`` column (renamed from ``name`` if necessary).

    Args:
        stat_view: One of the keys in ``STAT_VIEW_OPTIONS`` / ``STAT_VIEW_PARAMS``
                   (e.g. ``"S_S_2026"``, ``"S_L7"``).
        season: Fallback season year when not encoded in ``stat_view``.

    Returns:
        DataFrame with player metadata and stats columns for the requested view.
        Returns an empty DataFrame if the player pool is unavailable.
    """
    params = STAT_VIEW_PARAMS.get(stat_view, {"type": "total", "days": None, "season": None})
    view_type = params.get("type", "total")
    days = params.get("days")
    view_season = params.get("season") or season

    # Load base player pool (includes metadata + blended/ROS projections)
    try:
        pool = load_player_pool()
    except Exception:
        logger.exception("load_databank: failed to load player pool")
        return pd.DataFrame()

    if pool is None or pool.empty:
        return pd.DataFrame()

    # Normalise "name" → "player_name" (database returns "name"; pool may
    # already have "player_name" if enriched by _enrich_pool).
    if "player_name" not in pool.columns and "name" in pool.columns:
        pool = pool.rename(columns={"name": "player_name"})

    # Deduplicate by player_id — load_player_pool() can return multiple rows
    # per player when multiple projection systems exist.  Keep the first row
    # (which is the preferred/blended projection).
    if "player_id" in pool.columns:
        pool = pool.drop_duplicates(subset=["player_id"], keep="first")

    # Enrich with opponent, adds/drops (uses external APIs — graceful fallback)
    _enrich_databank_columns(pool)

    # For projection / special / advanced views, return pool as-is.
    # The pool already holds the best available projection stats.
    if view_type in ("proj", "special", "ranks", "research", "matchups", "opponents", "live", "advanced"):
        return pool.copy()

    # For total / avg / stddev views, compute from game logs.
    # "total" = season or rolling window sums (actual stats, not projections)
    # "avg"   = per-game averages over the window
    # "stddev"= standard deviation over the window
    if view_type in ("total", "avg", "stddev"):
        player_ids: list[int] = pool["player_id"].dropna().astype(int).tolist()
        if not player_ids:
            return pool.copy()

        rolled = compute_rolling_stats(
            player_ids,
            days=days,
            stat_type=view_type,
            season=view_season,
        )
        if rolled.empty:
            # No game log data — return pool metadata without stat columns
            return pool.copy()

        # Merge rolled stats onto pool metadata
        meta_cols = [c for c in pool.columns if c not in rolled.columns or c == "player_id"]
        merged = pool[meta_cols].merge(rolled, on="player_id", how="left")
        return merged

    # Fallback: return pool unchanged
    return pool.copy()


# ── Databank filtering ────────────────────────────────────────────────────────


def filter_databank(
    df: pd.DataFrame,
    position: str = "B",
    status: str = "A",
    mlb_team: str = "ALL",
    fantasy_team: str = "NONE",
    search: str = "",
    show_my_team: bool = False,
) -> pd.DataFrame:
    """Filter a databank DataFrame by common UI criteria.

    Args:
        df: DataFrame returned by ``load_databank()``.
        position: ``"B"`` = batters (``is_hitter==1``), ``"P"`` = pitchers
                  (``is_hitter==0``), or a specific position string (e.g.
                  ``"SS"``) matched via case-insensitive substring in the
                  ``positions`` column.  Anything else is treated as ``"ALL"``
                  (no filter).
        status: ``"A"`` = available (not on a fantasy roster), ``"FA"`` = free
                agents, ``"T"`` = taken (on a roster), ``"ALL"`` = no filter.
                Requires an ``is_available`` or ``roster_team`` column; skipped
                gracefully when absent.
        mlb_team: Filter to players on this MLB team.  ``"ALL"`` disables the
                  filter.  Matched against the ``team`` column.
        fantasy_team: Filter to players on this fantasy roster.  ``"NONE"``
                      disables the filter.  Matched against ``roster_team``.
        search: Case-insensitive substring search applied to ``player_name``
                (or ``name``).  Empty string disables the filter.
        show_my_team: When ``True`` and an ``is_user_team`` column is present,
                      restrict to rows where ``is_user_team`` is truthy.

    Returns:
        Filtered copy of ``df``.  Never modifies the input DataFrame.
    """
    result = df.copy()

    # ── Position filter ──────────────────────────────────────────────────────
    if position == "B":
        if "is_hitter" in result.columns:
            result = result[result["is_hitter"] == 1]
    elif position == "P":
        if "is_hitter" in result.columns:
            result = result[result["is_hitter"] == 0]
    elif position not in ("ALL", "", None):
        # Specific positional filter (e.g. "SS", "2B", "SP", "RP")
        if "positions" in result.columns:
            pos_upper = str(position).upper()
            result = result[result["positions"].fillna("").str.upper().str.contains(pos_upper, regex=False)]

    # ── Status filter ────────────────────────────────────────────────────────
    # Skip status filter when a specific fantasy team is selected — the user
    # wants to see that team's players regardless of availability.
    if not (fantasy_team and fantasy_team != "NONE"):
        if status == "A":
            # Available = not on any fantasy roster
            if "is_available" in result.columns:
                result = result[result["is_available"].fillna(True).astype(bool)]
            elif "roster_team" in result.columns:
                result = result[result["roster_team"].isna() | (result["roster_team"] == "")]
        elif status == "FA":
            if "roster_team" in result.columns:
                result = result[result["roster_team"].isna() | (result["roster_team"] == "")]
        elif status == "T":
            if "roster_team" in result.columns:
                result = result[result["roster_team"].notna() & (result["roster_team"] != "")]
        # "ALL" — no filter

    # ── MLB team filter ──────────────────────────────────────────────────────
    if mlb_team and mlb_team != "ALL":
        if "team" in result.columns:
            result = result[result["team"] == mlb_team]

    # ── Fantasy team filter ──────────────────────────────────────────────────
    if fantasy_team and fantasy_team != "NONE":
        if "roster_team" in result.columns:
            result = result[result["roster_team"] == fantasy_team]

    # ── Search filter ────────────────────────────────────────────────────────
    if search:
        name_col = "player_name" if "player_name" in result.columns else "name"
        if name_col in result.columns:
            result = result[result[name_col].fillna("").str.contains(search, case=False, regex=False)]
        else:
            # No name column — return empty
            result = result.iloc[0:0]

    # ── My team filter ───────────────────────────────────────────────────────
    if show_my_team and "is_user_team" in result.columns:
        result = result[result["is_user_team"].fillna(False).astype(bool)]

    return result


# ── HTML table rendering ──────────────────────────────────────────────────────

# Computed rate-stat column aliases: maps calc column → display column name
_CALC_ALIAS: dict[str, str] = {
    "avg_calc": "avg",
    "obp_calc": "obp",
    "era_calc": "era",
    "whip_calc": "whip",
}

# Rate-stat column names (all variants)
_RATE_STAT_COLS: set[str] = {"avg", "obp", "era", "whip", "avg_calc", "obp_calc", "era_calc", "whip_calc"}


def _format_cell(value: object, col: str) -> str:
    """Format a single cell value based on stat column type.

    Args:
        value: Raw cell value from the DataFrame.
        col: Column name (lowercase).

    Returns:
        Formatted string for display.
    """
    # NaN / None → dash
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return "-"
    except (TypeError, ValueError):
        pass

    col_lower = col.lower()

    # AVG / OBP style: .3f
    if col_lower in ("avg", "obp", "avg_calc", "obp_calc"):
        try:
            return f"{float(value):.3f}"
        except (TypeError, ValueError):
            return str(value)

    # ERA / WHIP style: .2f
    if col_lower in ("era", "whip", "era_calc", "whip_calc"):
        try:
            return f"{float(value):.2f}"
        except (TypeError, ValueError):
            return str(value)

    # IP: .1f
    if col_lower == "ip":
        try:
            return f"{float(value):.1f}"
        except (TypeError, ValueError):
            return str(value)

    # % Rostered: display as percentage
    if col_lower == "percent_owned":
        try:
            v = float(value)
            return f"{v:.1f}%" if v > 0 else "-"
        except (TypeError, ValueError):
            return "-"

    # ADP (Pre-Season rank): integer, hide if unranked
    if col_lower == "adp":
        try:
            v = float(value)
            return str(int(v)) if 0 < v < 999 else "-"
        except (TypeError, ValueError):
            return "-"

    # Consensus rank (Current): integer, hide if unranked
    if col_lower == "consensus_rank":
        try:
            v = float(value)
            return str(int(v)) if 0 < v < 9999 else "-"
        except (TypeError, ValueError):
            return "-"

    # Integer floats (e.g. 5.0 → "5")
    try:
        fval = float(value)
        if fval == int(fval):
            return str(int(fval))
        return str(fval)
    except (TypeError, ValueError):
        return str(value)


def render_databank_table(
    df: pd.DataFrame,
    stat_view: str = "S_S_2026",
    sort_col: str | None = None,
    sort_dir: str = "desc",
    is_pitcher: bool = False,
) -> str:
    """Render a player databank DataFrame as a sortable HTML table.

    Args:
        df: Player DataFrame from ``load_databank`` / ``filter_databank``.
        stat_view: Key from ``STAT_VIEW_OPTIONS`` used for the stat-group header.
        sort_col: Column to pre-sort by (None = no pre-sort).
        sort_dir: "asc" or "desc" initial sort direction.
        is_pitcher: True = pitcher stat columns; False = batter stat columns.

    Returns:
        Self-contained HTML string with embedded CSS and JavaScript.
    """
    if df is None or df.empty:
        return (
            f'<div style="padding:20px;text-align:center;color:{T["tx2"]};'
            f'font-family:Inter,sans-serif;">No players match the current filters.</div>'
        )

    # ── Column definitions ───────────────────────────────────────────────────
    if is_pitcher:
        stat_cols_raw = ["ip", "w", "l", "sv", "k", "era", "whip"]
        stat_labels = ["IP", "W", "L", "SV", "K", "ERA", "WHIP"]
    else:
        stat_cols_raw = ["r", "hr", "rbi", "sb", "avg", "obp"]
        stat_labels = ["R", "HR", "RBI", "SB", "AVG", "OBP"]

    # Resolve computed rate-stat aliases: prefer calc column, fall back to plain
    def _resolve_col(col: str) -> str:
        """Return the actual column present in df for a given stat col name."""
        calc = col + "_calc"
        if calc in df.columns:
            return calc
        return col

    stat_cols: list[str] = [_resolve_col(c) for c in stat_cols_raw]

    # Meta columns — core identity
    player_label = "Pitchers" if is_pitcher else "Batters"
    meta_cols: list[str] = ["player_name", "team", "positions"]
    meta_labels: list[str] = [player_label, "Team", "Pos"]

    # Roster status column
    if "roster_team" in df.columns:
        df = df.copy()
        df["_roster_status"] = df["roster_team"].fillna("FA").replace("", "FA")
        meta_cols.append("_roster_status")
        meta_labels.append("Roster Status")

    # Optional GP column (try multiple names)
    gp_col: str | None = None
    for candidate in ("games_played", "gp", "ytd_gp"):
        if candidate in df.columns:
            gp_col = candidate
            break
    if gp_col is not None:
        meta_cols.append(gp_col)
        meta_labels.append("GP*")

    # Opponent column (today's matchup)
    if "opponent" in df.columns:
        meta_cols.append("opponent")
        _now = datetime.now(UTC)
        meta_labels.append(f"Opp: {_now.month}/{_now.day}")

    # Pre-Season ranking (ADP)
    if "adp" in df.columns:
        meta_cols.append("adp")
        meta_labels.append("Pre-Season")

    # Current ranking (ECR consensus)
    if "consensus_rank" in df.columns:
        meta_cols.append("consensus_rank")
        meta_labels.append("Current")

    # Adds / Drops
    if "adds" in df.columns:
        meta_cols.append("adds")
        meta_labels.append("Adds")
    if "drops" in df.columns:
        meta_cols.append("drops")
        meta_labels.append("Drops")

    # % Rostered
    if "percent_owned" in df.columns:
        meta_cols.append("percent_owned")
        meta_labels.append("% Ros")

    # Build full column / label lists
    all_cols: list[str] = meta_cols.copy()
    all_labels: list[str] = meta_labels.copy()
    all_cols.extend(stat_cols)
    all_labels.extend(stat_labels)

    n_meta = len(meta_cols)
    n_stat = len(stat_cols)

    # Stat group label from STAT_VIEW_OPTIONS
    group_label: str = STAT_VIEW_OPTIONS.get(stat_view, stat_view)

    # ── CSS ──────────────────────────────────────────────────────────────────
    css = f"""
<style>
.hdb-table-wrap {{
    overflow-x: auto;
    border-radius: 8px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.10);
    background: {T["card"]};
}}
.hdb-table {{
    width: 100%;
    border-collapse: collapse;
    font-family: Inter, -apple-system, BlinkMacSystemFont, sans-serif;
    font-size: 13px;
}}
.hdb-table thead th {{
    background: linear-gradient(135deg, #16213e, #1a1a2e);
    color: #ffffff;
    padding: 8px 10px;
    position: sticky;
    top: 0;
    cursor: pointer;
    white-space: nowrap;
    text-align: left;
    user-select: none;
}}
.hdb-table thead th:hover {{
    background: #1e2a45;
}}
.sort-arrow {{
    font-size: 10px;
    margin-left: 3px;
    opacity: 0.5;
}}
.hdb-table thead th.sorted .sort-arrow {{
    opacity: 1.0;
    color: #ff6d00;
}}
.hdb-table .stat-group {{
    background: linear-gradient(135deg, #1a1a2e, #16213e);
    color: #ff6d00;
    text-align: center;
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.04em;
    padding: 4px 8px;
}}
.hdb-table tbody tr {{
    border-bottom: 1px solid #e5e7eb;
}}
.hdb-table tbody tr:nth-child(even) {{
    background: #fafaf8;
}}
.hdb-table tbody tr:hover {{
    background: #fff7ed !important;
}}
.hdb-table td {{
    padding: 8px 10px;
    white-space: nowrap;
}}
.hdb-table .player-name {{
    font-weight: 600;
    color: {T["sky"]};
    min-width: 160px;
}}
.hdb-table .stat-cell {{
    text-align: right;
    font-variant-numeric: tabular-nums;
}}
.hdb-table .rate-stat {{
    font-weight: 600;
}}
</style>
"""

    # ── JavaScript ───────────────────────────────────────────────────────────
    # Arrow characters as Python strings to avoid JS unicode escapes
    up_arrow = "\u25b2"
    down_arrow = "\u25bc"
    js = f"""
<script>
function sortTable(colIdx, headerEl) {{
    var table = headerEl.closest('table');
    var tbody = table.querySelector('tbody');
    var rows = Array.from(tbody.querySelectorAll('tr'));
    var headers = table.querySelectorAll('thead tr:last-child th');

    // Determine sort direction: toggle if already sorted, else default desc
    var asc = true;
    if (headerEl.classList.contains('sorted')) {{
        asc = headerEl.dataset.sortDir !== 'asc';
    }}

    // Clear sorted state on all headers
    headers.forEach(function(h) {{
        h.classList.remove('sorted');
        h.dataset.sortDir = '';
        var arrow = h.querySelector('.sort-arrow');
        if (arrow) {{ arrow.textContent = '{up_arrow}'; }}
    }});

    // Mark this header as sorted
    headerEl.classList.add('sorted');
    headerEl.dataset.sortDir = asc ? 'asc' : 'desc';
    var thisArrow = headerEl.querySelector('.sort-arrow');
    if (thisArrow) {{ thisArrow.textContent = asc ? '{up_arrow}' : '{down_arrow}'; }}

    // Sort rows by data-sort attribute
    rows.sort(function(a, b) {{
        var aCells = a.querySelectorAll('td');
        var bCells = b.querySelectorAll('td');
        var aVal = aCells[colIdx] ? aCells[colIdx].dataset.sort : '';
        var bVal = bCells[colIdx] ? bCells[colIdx].dataset.sort : '';

        var aNum = parseFloat(aVal);
        var bNum = parseFloat(bVal);
        var useNum = !isNaN(aNum) && !isNaN(bNum);

        var cmp = useNum ? (aNum - bNum) : aVal.localeCompare(bVal);
        return asc ? cmp : -cmp;
    }});

    // Re-append sorted rows to tbody
    rows.forEach(function(r) {{ tbody.appendChild(r); }});
}}
</script>
"""

    # ── Table HTML ───────────────────────────────────────────────────────────
    # Stat group header row
    group_row = "<tr>"
    for _ in range(n_meta):
        group_row += (
            '<th class="stat-group" style="background:linear-gradient(135deg, #16213e, #1a1a2e);color:#ffffff;"></th>'
        )
    group_row += f'<th class="stat-group" colspan="{n_stat}">{_html_escape(group_label)}</th>'
    group_row += "</tr>"

    # Column header row (with onclick for JS sort)
    header_row = "<tr>"
    for idx, (col, label) in enumerate(zip(all_cols, all_labels)):
        header_row += (
            f'<th onclick="sortTable({idx}, this)">{_html_escape(label)}<span class="sort-arrow">{up_arrow}</span></th>'
        )
    header_row += "</tr>"

    # Data rows
    data_rows = ""
    for _, row in df.iterrows():
        data_rows += "<tr>"
        for col in all_cols:
            raw_val = row.get(col, None) if hasattr(row, "get") else getattr(row, col, None)
            formatted = _format_cell(raw_val, col)

            # data-sort: raw numeric string for JS comparison, empty for missing
            try:
                if raw_val is None or (isinstance(raw_val, float) and math.isnan(raw_val)):
                    sort_val = ""
                else:
                    sort_val = str(float(raw_val))
            except (TypeError, ValueError):
                sort_val = str(raw_val) if raw_val is not None else ""

            # CSS classes
            is_rate = col.lower() in _RATE_STAT_COLS
            if col == "player_name":
                td_class = "player-name"
            elif col in ("team", "positions") or col == gp_col:
                td_class = ""
            else:
                td_class = "stat-cell rate-stat" if is_rate else "stat-cell"

            class_attr = f' class="{td_class}"' if td_class else ""
            data_rows += f'<td{class_attr} data-sort="{_html_escape(sort_val)}">{_html_escape(formatted)}</td>'
        data_rows += "</tr>"

    table_html = (
        f'<div class="hdb-table-wrap">'
        f'<table class="hdb-table">'
        f"<thead>{group_row}{header_row}</thead>"
        f"<tbody>{data_rows}</tbody>"
        f"</table>"
        f"</div>"
    )

    # Wrap in scrollable container
    return f'<div style="max-height:600px;overflow-y:auto;">{css}{table_html}{js}</div>'


def _html_escape(text: str) -> str:
    """Escape HTML special characters in a string.

    Args:
        text: Input string.

    Returns:
        HTML-safe string with &, <, >, ", ' escaped.
    """
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


_EST_OFFSET = timedelta(hours=-5)  # EST = UTC-5 (no DST adjustment)


def _parse_iso_timestamp(ts_str: str) -> datetime | None:
    """Parse an ISO-format timestamp string, returning a datetime or None."""
    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(ts_str, fmt)
        except ValueError:
            continue
    return None


def get_data_as_of_label(stat_view: str, season: int = 2026) -> str:
    """Return a human-readable 'as of' label for the current stat view.

    For total/avg/stddev views, returns the most recent game date from
    game_logs (e.g. "As of 04/11's games").  For projection views, returns
    "Based on ROS projections".  Returns empty string if no data available.

    Args:
        stat_view: Key from ``STAT_VIEW_OPTIONS``.
        season: Default season year.

    Returns:
        Label string for display, or empty string.
    """
    params = STAT_VIEW_PARAMS.get(stat_view, {})
    view_type = params.get("type", "total")
    view_season = params.get("season") or season
    days = params.get("days")

    if view_type in ("proj",):
        return "Based on ROS projections"

    if view_type in ("total", "avg", "stddev"):
        conn = get_connection()
        try:
            # Try game_logs first for per-game granularity
            clauses: list[str] = []
            sql_params: list = []
            if view_season:
                clauses.append("season = ?")
                sql_params.append(view_season)
            if days is not None and days > 0:
                cutoff = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%d")
                clauses.append("game_date >= ?")
                sql_params.append(cutoff)
            where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
            cursor = conn.cursor()
            cursor.execute(f"SELECT MAX(game_date) FROM game_logs {where}", sql_params)
            row = cursor.fetchone()
            if row and row[0]:
                dt = datetime.strptime(row[0], "%Y-%m-%d")
                return f"As of {dt.strftime('%m/%d')}'s games"

            # Fallback: use refresh_log timestamp for season_stats.
            # MLB season stats reflect completed games. The refresh runs
            # during the day, so stats are through the previous day's
            # completed games (the refresh date minus 1 day).
            from src.database import get_refresh_status

            for src in ("season_stats", "game_logs"):
                rs = get_refresh_status(src)
                if rs and rs.get("last_refresh"):
                    ts = str(rs["last_refresh"])
                    dt = _parse_iso_timestamp(ts)
                    if dt is not None:
                        # Stats are through the day before the refresh
                        stats_through = dt - timedelta(days=1)
                        return f"As of {stats_through.month}/{stats_through.day}'s games"
        except Exception:
            logger.debug("Failed to get latest game date for stat view %s", stat_view)
        finally:
            conn.close()
        return ""

    if view_type == "advanced":
        return "Statcast advanced metrics"

    return ""


def get_data_refreshed_label(stat_view: str) -> str:
    """Return a 'Refreshed as of ...' label based on the refresh log.

    Checks the most relevant refresh source for the stat view (live_stats
    for season/rolling views, projections for projection views) and formats
    the timestamp in local time.

    Returns empty string if no refresh record found.
    """
    from src.database import get_refresh_status

    params = STAT_VIEW_PARAMS.get(stat_view, {})
    view_type = params.get("type", "total")

    # Pick the most relevant refresh sources (try in order)
    if view_type == "proj":
        sources = ["ros_projections", "projections"]
    elif view_type == "advanced":
        sources = ["batting_stats", "statcast"]
    else:
        sources = ["season_stats", "game_logs", "live_stats"]

    for source in sources:
        status = get_refresh_status(source)
        if status and status.get("last_refresh"):
            dt = _parse_iso_timestamp(str(status["last_refresh"]))
            if dt is not None:
                # Convert UTC → EST (UTC-5) for display
                est_dt = dt + _EST_OFFSET
                time_str = est_dt.strftime("%I:%M %p").lstrip("0")
                return f"Refreshed {est_dt.month}/{est_dt.day}, {time_str} EST"
    return ""


def export_to_excel(df: pd.DataFrame, stat_view_label: str) -> bytes:
    """Export the databank DataFrame to a branded Excel file.

    Args:
        df: Filtered and sorted DataFrame to export.
        stat_view_label: Human-readable stat view name for sheet name.

    Returns:
        Excel file as bytes.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    sheet_name = stat_view_label[:31]  # Excel 31-char limit

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = writer.book
        ws = wb[sheet_name]

        # HEATER branding
        flame_fill = PatternFill(start_color="E63946", end_color="E63946", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        stat_font = Font(name="Calibri", size=10)

        # Style header row
        for cell in ws[1]:
            cell.fill = flame_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-width columns
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if len(df) > 0 else 0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 20)

        # Format stat cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column <= len(df.columns):
                    col_name = df.columns[cell.column - 1]
                    cell.font = stat_font
                    if col_name in ("avg", "obp", "avg_calc", "obp_calc"):
                        cell.number_format = "0.000"
                    elif col_name in ("era", "whip", "era_calc", "whip_calc"):
                        cell.number_format = "0.00"
                    elif col_name == "ip":
                        cell.number_format = "0.0"

    output.seek(0)
    return output.getvalue()
